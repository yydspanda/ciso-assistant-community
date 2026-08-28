"""Adversarial IAM coverage for compliance action-plan exports."""

from __future__ import annotations

import csv
import io
import uuid

import pytest
from openpyxl import load_workbook

from core.models import (
    Actor,
    AppliedControl,
    ComplianceAssessment,
    EvidenceRevision,
    Framework,
    ReferenceControl,
)
from core.serializers import ComplianceAssessmentActionPlanSerializer
from core.tests.test_compliance_assessment_tree_iam import (
    _client,
    _grant,
    audit_iam_world as _audit_iam_world_fixture,
)
from iam.models import User


pytestmark = pytest.mark.django_db
audit_iam_world = _audit_iam_world_fixture

EXPORT_ACTIONS = ("action_plan_csv", "action_plan_xlsx", "action_plan_pdf")
SUPPLEMENTAL_EXPORT_PERMISSIONS = {
    "view_actor",
    "view_evidencerevision",
    "view_entity",
    "view_folder",
    "view_perimeter",
    "view_requirementnode",
    "view_team",
    "view_user",
}


def _export_url(target, action: str) -> str:
    return f"/api/compliance-assessments/{target.id}/{action}/"


def _make_user(email_prefix: str, folder):
    user = User.objects.create_user(
        f"{email_prefix}-{uuid.uuid4().hex}@action-plan-iam.tests"
    )
    user.folder = folder
    user.save(update_fields=["folder"])
    return user


@pytest.fixture
def action_plan_iam_world(audit_iam_world):
    world = audit_iam_world
    child = world["child_folder"]
    auditor = world["auditor"]
    _grant(
        auditor,
        f"Action-plan related reader {uuid.uuid4().hex}",
        SUPPLEMENTAL_EXPORT_PERMISSIONS,
        child,
    )

    reference_control = ReferenceControl.objects.create(
        name="Action-plan reference control",
        ref_id="AP-REF",
        urn=f"urn:test:action-plan-reference:{uuid.uuid4().hex}",
        folder=child,
    )
    control = world["visible_control"]
    AppliedControl.objects.filter(id=control.id).update(
        reference_control=reference_control,
        description="Visible action-plan description",
    )
    control.reference_control = reference_control
    control.description = "Visible action-plan description"
    world["assigned_ra"].applied_controls.add(control)

    evidence = world["visible_evidence"]
    control.evidences.add(evidence)
    visible_revision = EvidenceRevision.objects.create(
        evidence=evidence,
        folder=child,
        version=1,
        observation="visible revision",
    )

    owner_user = _make_user("visible-owner", child)
    owner = Actor.objects.get(user=owner_user)
    control.owner.add(owner)
    evidence.owner.add(owner)

    return {
        **world,
        "control": control,
        "evidence": evidence,
        "visible_revision": visible_revision,
        "owner": owner,
        "reference_control": reference_control,
    }


def test_fully_authorized_exports_keep_legacy_wire_formats(
    action_plan_iam_world, monkeypatch
):
    world = action_plan_iam_world
    target = world["target"]
    client = _client(world["auditor"])

    csv_response = client.get(_export_url(target, "action_plan_csv"))
    assert csv_response.status_code == 200, csv_response.content
    assert csv_response["Content-Type"] == "text/csv; charset=utf-8"
    csv_rows = list(csv.reader(io.StringIO(csv_response.content.decode("utf-8-sig"))))
    assert csv_rows[0] == [
        "Name",
        "Description",
        "Category",
        "CSF Function",
        "Priority",
        "Status",
        "ETA",
        "Expiry date",
        "Effort",
        "Impact",
        "Cost",
        "Covered requirements",
        "Associated evidences",
        "Evidence attachments",
    ]
    assert csv_rows[1][0] == world["control"].name
    assert world["assigned_requirement"].ref_id in csv_rows[1][11]
    assert world["evidence"].name in csv_rows[1][12]

    xlsx_response = client.get(_export_url(target, "action_plan_xlsx"))
    assert xlsx_response.status_code == 200, xlsx_response.content
    assert xlsx_response["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(io.BytesIO(xlsx_response.content), read_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[0] == (
        "name",
        "description",
        "category",
        "csf_function",
        "priority",
        "status",
        "eta",
        "expiry_date",
        "effort",
        "impact",
        "cost",
        "covered_requirements",
        "associated_evidences",
        "evidence_attachments",
    )
    assert rows[1][0] == world["control"].name

    rendered = {}

    class FakeHTML:
        def __init__(self, *, string):
            rendered["html"] = string

        def write_pdf(self):
            return b"%PDF-action-plan-iam"

    monkeypatch.setattr("core.views.HTML", FakeHTML)
    pdf_response = client.get(_export_url(target, "action_plan_pdf"))
    assert pdf_response.status_code == 200, pdf_response.content
    assert pdf_response["Content-Type"] == "application/pdf"
    assert pdf_response.content == b"%PDF-action-plan-iam"
    assert target.name in rendered["html"]
    assert target.perimeter.name in rendered["html"]
    assert world["assigned_requirement"].ref_id in rendered["html"]


@pytest.mark.parametrize(
    "hidden_relation",
    (
        "requirement_assessment",
        "applied_control",
        "evidence",
        "evidence_revision",
        "actor",
        "reference_control",
    ),
)
def test_exports_fail_closed_instead_of_partially_omitting_hidden_relations(
    action_plan_iam_world, hidden_relation
):
    world = action_plan_iam_world
    hidden = world["hidden_folder"]

    if hidden_relation == "requirement_assessment":
        world["assigned_ra"].folder = hidden
        world["assigned_ra"].save(update_fields=["folder"])
    elif hidden_relation == "applied_control":
        world["assigned_ra"].applied_controls.add(world["hidden_control"])
    elif hidden_relation == "evidence":
        world["control"].evidences.add(world["hidden_evidence"])
    elif hidden_relation == "evidence_revision":
        hidden_revision = EvidenceRevision.objects.create(
            evidence=world["evidence"],
            folder=hidden,
            version=2,
            observation="hidden revision",
        )
        # EvidenceRevision.save() aligns the folder to its Evidence.  Move it
        # after creation to construct the adversarial legacy/corrupt relation
        # this read boundary must still reject.
        EvidenceRevision.objects.filter(id=hidden_revision.id).update(folder=hidden)
    elif hidden_relation == "actor":
        hidden_owner_user = _make_user("hidden-owner", hidden)
        world["control"].owner.add(Actor.objects.get(user=hidden_owner_user))
    elif hidden_relation == "reference_control":
        hidden_reference = ReferenceControl.objects.create(
            name="Hidden action-plan reference",
            ref_id="AP-REF-HIDDEN",
            urn=f"urn:test:action-plan-reference:hidden:{uuid.uuid4().hex}",
            folder=hidden,
        )
        AppliedControl.objects.filter(id=world["control"].id).update(
            reference_control=hidden_reference
        )
    else:  # pragma: no cover - parameter contract
        raise AssertionError(hidden_relation)

    client = _client(world["auditor"])
    for action in EXPORT_ACTIONS:
        response = client.get(_export_url(world["target"], action))
        assert response.status_code == 403, (action, response.content)


@pytest.mark.parametrize("hidden_field", ("applied_controls", "evidences"))
def test_exports_require_both_action_plan_fields_to_be_auditor_visible(
    action_plan_iam_world, hidden_field
):
    world = action_plan_iam_world
    target = world["target"]
    target.field_visibility = {
        **target.field_visibility,
        hidden_field: {"auditor": "hidden", "respondent": "hidden"},
    }
    target.save(update_fields=["field_visibility"])

    client = _client(world["auditor"])
    for action in EXPORT_ACTIONS:
        assert client.get(_export_url(target, action)).status_code == 403


def test_ordinary_audit_reader_cannot_reach_any_action_plan_export(
    action_plan_iam_world,
):
    world = action_plan_iam_world
    client = _client(world["respondent"])

    for action in EXPORT_ACTIONS:
        assert client.get(_export_url(world["target"], action)).status_code == 403


@pytest.mark.parametrize("missing_permission", ("view_folder", "view_perimeter"))
def test_pdf_independently_authorizes_folder_and_perimeter(
    action_plan_iam_world, missing_permission
):
    world = action_plan_iam_world
    user = _make_user("pdf-metadata-reader", world["child_folder"])
    permissions = {
        "view_actor",
        "view_appliedcontrol",
        "view_complianceassessment",
        "view_compliance_assessment_full",
        "view_entity",
        "view_evidence",
        "view_evidencerevision",
        "view_folder",
        "view_framework",
        "view_perimeter",
        "view_referencecontrol",
        "view_requirementassessment",
        "view_requirementnode",
        "view_securityexception",
        "view_team",
        "view_user",
    }
    permissions.remove(missing_permission)
    _grant(
        user,
        f"PDF metadata reader {uuid.uuid4().hex}",
        permissions,
        world["child_folder"],
    )
    _grant(
        user,
        f"PDF framework-node reader {uuid.uuid4().hex}",
        {"view_requirementnode"},
        world["child_folder"].get_root_folder(),
        is_recursive=False,
    )
    client = _client(user)

    # The CSV has no folder/perimeter column and remains available. The PDF
    # prints both labels, so each requires its own object-level read grant.
    assert (
        client.get(_export_url(world["target"], "action_plan_csv")).status_code == 200
    )
    list_response = client.get(
        f"/api/compliance-assessments/{world['target'].id}/action-plan/"
    )
    budget_response = client.get(
        f"/api/compliance-assessments/{world['target'].id}/action-plan/budget-overview/"
    )
    expected_non_pdf_status = 403 if missing_permission == "view_folder" else 200
    assert list_response.status_code == expected_non_pdf_status
    assert budget_response.status_code == expected_non_pdf_status
    assert (
        client.get(_export_url(world["target"], "action_plan_pdf")).status_code == 403
    )


def test_list_and_budget_use_narrow_filters_and_omit_risk_derived_ranking(
    action_plan_iam_world,
):
    world = action_plan_iam_world
    client = _client(world["auditor"])
    list_url = f"/api/compliance-assessments/{world['target'].id}/action-plan/"
    budget_url = f"{list_url}budget-overview/"

    baseline_list = client.get(list_url)
    baseline_budget = client.get(budget_url)
    assert baseline_list.status_code == 200, baseline_list.content
    assert baseline_budget.status_code == 200, baseline_budget.content
    rows = baseline_list.json().get("results", baseline_list.json())
    assert rows
    assert "ranking_score" not in rows[0]

    hidden_id = str(uuid.uuid4())
    unsafe_params = (
        {"linked_models": "risk_scenarios"},
        {"risk_scenarios": hidden_id},
    )
    for params in unsafe_params:
        filtered_list = client.get(list_url, params)
        filtered_budget = client.get(budget_url, params)
        assert filtered_list.status_code == 200
        assert filtered_budget.status_code == 200
        assert filtered_list.json() == baseline_list.json()
        assert filtered_budget.json() == baseline_budget.json()


def test_empty_assessment_visibility_uses_hidden_framework_policy(
    action_plan_iam_world,
):
    world = action_plan_iam_world
    Framework.objects.filter(id=world["target"].framework_id).update(
        field_visibility={
            "applied_controls": {"auditor": "hidden", "respondent": "hidden"},
            "evidences": {"auditor": "hidden", "respondent": "hidden"},
        }
    )
    ComplianceAssessment.objects.filter(id=world["target"].id).update(
        field_visibility={}
    )
    client = _client(world["auditor"])

    urls = [_export_url(world["target"], action) for action in EXPORT_ACTIONS] + [
        f"/api/compliance-assessments/{world['target'].id}/action-plan/",
        f"/api/compliance-assessments/{world['target'].id}/action-plan/budget-overview/",
    ]
    for url in urls:
        assert client.get(url).status_code == 403


@pytest.mark.parametrize("mutation", ("unlink", "hide-policy"))
def test_terminal_reproof_is_bound_to_the_serialized_projection(
    action_plan_iam_world, monkeypatch, mutation
):
    world = action_plan_iam_world
    original = ComplianceAssessmentActionPlanSerializer.to_representation
    mutated = False

    def mutate_after_serialization(serializer, instance):
        nonlocal mutated
        payload = original(serializer, instance)
        if not mutated:
            mutated = True
            if mutation == "unlink":
                instance.evidences.clear()
            else:
                visibility = {
                    key: dict(value)
                    for key, value in world["target"].field_visibility.items()
                }
                visibility["evidences"] = {
                    "auditor": "hidden",
                    "respondent": "hidden",
                }
                ComplianceAssessment.objects.filter(id=world["target"].id).update(
                    field_visibility=visibility
                )
        return payload

    monkeypatch.setattr(
        ComplianceAssessmentActionPlanSerializer,
        "to_representation",
        mutate_after_serialization,
    )
    response = _client(world["auditor"]).get(
        _export_url(world["target"], "action_plan_csv")
    )

    assert mutated is True
    assert response.status_code == 403


def test_action_plan_serializer_does_not_query_requirements_without_request_context(
    action_plan_iam_world,
):
    world = action_plan_iam_world
    data = ComplianceAssessmentActionPlanSerializer(
        world["control"], context={"pk": world["target"].id}
    ).data

    assert data["requirement_assessments"] == []
